#!/usr/bin/env python
# -*- coding: utf-8 -*-
import sys

if sys.version_info[0] != 2 or sys.version_info[1] < 7:
    sys.exit("This script requires Python version 2.7; you're runnning " +
             ".".join(str(x) for x in sys.version_info))

from collections import OrderedDict
from zipfile import ZipFile
import argparse
import csv
import datetime
import dbf
import json
import logging
import os
import re
import urllib2
# https://pypi.python.org/pypi/dbf

club = 'dworp'
clubnummer = 228
# huidig seizoen
seizoen = "ni2122"
# ploegen = (1,) indien 1 ploeg in de interclubs
# ploegen = (1, 2) indien 2 ploegen in de interclubs
# ploegen = (1, 2, 3) indien 3 ploegen in de interclubs
ploegen = (1, 2, 3, 4)
# reeksnamen
reeksnamen = ('3b', '4a', '4c', '5k')
# url voor locatie bestanden kbsb
url_locatie = 'https://www.frbe-kbsb.be/sites/manager/ICN/21-22/'
# dbase-bestand van kbsb (zit in zip-file)
zip_input = 'Datas.zip'
dbf_input = 'Part24L.DBF'
# dictionary om scores uit dbase-bestand om te zetten in getallen
score = {'0': 0, '0F': 0, '1': 1, '1F': 1, '0.5': 0.5}
# excel-bestand van kbsb
xls_input = 'NationaalInterclub2021-2022_Uitslagen.xlsm'
# worksheets in excel-bestand
ws_ronden = ('R1', 'R2', 'R3', 'R4', 'R5', 'R6',
             'R7', 'R8', 'R9', 'R10', 'R11')
# cellen in excel-bestand, wijzigt elk seizoen indien reeks waarin
# ploegen spelen, wijzigt. Eerste veld is steeds voor Dworp 1, tweede
# voor Dworp 2, lege string indien geen Dworp 2
cell_ranking_begin = ('A47', 'A272', 'A317', 'A347')
cell_ranking_eind = ('Q59', 'Q284', 'Q329', 'Q359')
cell_uitslag_begin = ('A16', 'A56', 'A64', 'M64')
cell_uitslag_eind = ('E21', 'E61', 'E69', 'Q69')
cell_datum = 'H5'
# ni_dir is de directory voor alle data van seizoen
cur_dir = os.path.abspath(os.curdir)
ni_dir = os.path.join(cur_dir, "..", "_data", "ni", seizoen)
ni_dir = os.path.normpath(ni_dir)
try:
    os.makedirs(ni_dir)
except OSError:
    pass

app_name = os.path.basename(sys.argv[0])
app_name = os.path.splitext(app_name)[0]

if sys.platform == "win32":
    data_locatie = os.path.join(os.environ['LOCALAPPDATA'], app_name)
elif sys.platform == "darwin":
    data_locatie = os.path.join(os.path.expanduser(
        '/Library/Application Support'), app_name)
else:
    data_locatie = os.path.join("/var/lib", app_name)

log_locatie = "/var/log"
lib_locate = "/var/lib/update_ni"
log_file = 'update_ni.log'

verbeter_team_namen = [
    ('Dzd', 'DZD'),
    ('Kgsrl', 'KGSRL'),
    ('Creb', 'CREB'),
    ('Tsm', 'TSM'),
    ('O/D', 'o/d'),
    ('000 Bye 5I', 'Bye'),
    ]


def maak_csv_ranking(verbose=False):
    wb = load_workbook(xls_input, data_only=True)
    ws = wb['Ranking']
    for i, ploeg in enumerate(ploegen):
        cell_range = ws[cell_ranking_begin[i]:cell_ranking_eind[i]]
        string_list = []
        for row in cell_range:
            l = []
            for cell in row:
                if cell.value:
                    l.append(unicode(cell.value).lower().encode('utf-8'))
                elif cell.value == 0:
                    l.append('0')
                else:
                    l.append('')
            string_list.append(l)
        # schrijf csv-bestand
        bestandsnaam = os.path.join(
            ni_dir, 'tussenstand_{0}_{1}_{2}.csv'.format(club, ploeg, reeksnamen[i]))
        logger.debug(string_list)
        backup_vorig_bestand(bestandsnaam)
        with open(bestandsnaam, 'wb') as csvfile:
            csvwriter = csv.writer(csvfile, quoting=csv.QUOTE_NONE)
            csvwriter.writerows(string_list)


def maak_json_ploeg_uitslagen(verbose=False):
    wb = load_workbook(xls_input, data_only=True)
    for i, ploeg in enumerate(ploegen):
        l = []
        for n, worksheet in enumerate(ws_ronden):
            ws = wb[worksheet]
            try:
                datum = ws[cell_datum].value.strftime('%d-%m-%Y')
            except:
                # hack voor 11de ronde
                datum = ws[cell_datum].value[:10].replace('/', '-')
            ronde = OrderedDict(
                [("ronde", "ronde %d" % (n + 1,)), ("datum", datum)])
            matches = []
            cell_range = ws[cell_uitslag_begin[i]:cell_uitslag_eind[i]]
            for row in cell_range:
                team1 = row[0].value.title()
                team2 = row[1].value.title()
                for k, v in verbeter_team_namen:
                    team1 = team1.replace(k, v)
                    team2 = team2.replace(k, v)
                if row[2].value is not None:
                    team1_res = str(row[2].value)
                else:
                    team1_res = ''
                if row[4].value is not None:
                    team2_res = str(row[4].value)
                else:
                    team2_res = ''
                matches.append(OrderedDict(
                    [("tr", team1_res), ("ur", team2_res), ("thuis", team1),
                     ("uit", team2)]))
            ronde["uitslagen"] = matches
            l.append(ronde)
        json_string = json.dumps(l, indent=2, separators=(',', ': '))
        # remove a lot of whitespace and newlines with regular expression
        p = re.compile(r'''
            (\{)\s+
            ("tr":\s"[0-9.]*",)\s+
            ("ur":\s"[0-9.]*",)\s+
            ("thuis":\s".+",)\s+
            ("uit":\s".+")\s+
            (\})
            ''', re.VERBOSE)
        json_string = p.sub(r'\1\2 \3 \4 \5\6', json_string)
        logger.debug(json_string)
        # schrijf json-bestand
        bestandsnaam = os.path.join(
            ni_dir, 'ploegen_{0}_{1}_{2}.json'.format(club, ploeg, reeksnamen[i]))
        backup_vorig_bestand(bestandsnaam)
        with open(bestandsnaam, 'w') as f:
            f.write(json_string)


def maak_json_individuele_uitslagen(verbose=False):
    table = dbf.Table(dbf_input)
    table.open()

    w_index = table.create_index(lambda rec: rec.clb_icn_b)
    z_index = table.create_index(lambda rec: rec.clb_icn_n)

    w_records = w_index.search(match=(clubnummer,))
    z_records = z_index.search(match=(clubnummer,))
    records = w_records + z_records

    d = {}
    teams = set()
    for r in records:
        w_club = str(r.clb_icn_b) + ' ' + r.equipe_b.rstrip()
        z_club = str(r.clb_icn_n) + ' ' + r.equipe_n.rstrip()
        # 2 ploegen van 1 club kunnen in dezelfde reeks zitten
        if clubnummer == r.clb_icn_b:
            team = r.equipe_b.rstrip()
            teams.add(team)
            d[(team, r.ronde, r.tableau)] = (
                r.date, w_club, z_club, r.res.rstrip(), r.mat_b,
                r.nom_b.rstrip(), r.elo_b, r.mat_n, r.nom_n.rstrip(), r.elo_n)
        if clubnummer == r.clb_icn_n:
            team = r.equipe_n.rstrip()
            teams.add(team)
            d[(team, r.ronde, r.tableau)] = (
                r.date, w_club, z_club, r.res.rstrip(), r.mat_b,
                r.nom_b.rstrip(), r.elo_b, r.mat_n, r.nom_n.rstrip(), r.elo_n)

    teams = list(teams)
    teams.sort()
    for team in teams:
        l = []
        for r in range(1, 12):
            if d[(team, r, 1)][3] == '-':
                continue
            boards = []
            tscore, uscore, trating, urating, tdiv, udiv = (
                0, 0, 0, 0, 0.0, 0.0)
            for b in range(1, 9):
                try:
                    w_res, b_res = d[(team, r, b)][3].replace(
                        u'\u00bd', '0.5').split('-')
                    if b % 2:
                        boards.append(
                            OrderedDict([('tr', w_res), ('ur', b_res),
                                         ('tspeler', d[(team, r, b)][5]),
                                         ('uspeler', d[(team, r, b)][8]),
                                         ('telo', d[(team, r, b)][6]),
                                         ('uelo', d[(team, r, b)][9]),
                                         ('tstam', d[(team, r, b)][4]),
                                         ('ustam', d[(team, r, b)][7])]))
                        tscore += score[w_res]
                        uscore += score[b_res]
                        if d[(team, r, b)][6]:
                            trating += d[(team, r, b)][6]
                            tdiv += 1
                        if d[(team, r, b)][9]:
                            urating += d[(team, r, b)][9]
                            udiv += 1
                    else:
                        boards.append(
                            OrderedDict([('tr', b_res), ('ur', w_res),
                                         ('tspeler', d[(team, r, b)][8]),
                                         ('uspeler', d[(team, r, b)][5]),
                                         ('telo', d[(team, r, b)][9]),
                                         ('uelo', d[(team, r, b)][6]),
                                         ('tstam', d[(team, r, b)][7]),
                                         ('ustam', d[(team, r, b)][4])]))
                        tscore += score[b_res]
                        uscore += score[w_res]
                        if d[(team, r, b)][9]:
                            trating += d[(team, r, b)][9]
                            tdiv += 1
                        if d[(team, r, b)][6]:
                            urating += d[(team, r, b)][6]
                            udiv += 1
                except KeyError:
                    pass
            datum = d[(team, r, 1)][0].strftime('%d-%m-%Y')
            ronde = OrderedDict([("ronde", str(r)), ("datum", datum),
                                 ('thuis', d[(team, r, 1)][1]),
                                 ('uit', d[(team, r, 1)][2])])
            ronde["tscore"] = tscore
            ronde["uscore"] = uscore
            try:
                ronde["tgem"] = int(round(trating / tdiv))
            except:
                ronde["tgem"] = 0
            try:
                ronde["ugem"] = int(round(urating / udiv))
            except:
                ronde["ugem"] = 0
            ronde["borden"] = boards
            l.append(ronde)
        json_string = json.dumps(l, indent=2, separators=(',', ': '))
        # remove a lot of whitespace and newlines with regular expression
        p = re.compile(r'''
            (\{)\s+
            ("tr":\s"[015F.]+",)\s+
            ("ur":\s"[015F.]+",)\s+
            ("tspeler":\s".+",)\s+
            ("uspeler":\s".+",)\s+
            ("telo":\s\d+,)\s+
            ("uelo":\s\d+,)\s+
            ("tstam":\s\d+,)\s+
            ("ustam":\s\d+)\s+
            (\})
            ''', re.VERBOSE)
        json_string = p.sub(r'\1\2 \3 \4 \5 \6 \7 \8 \9\10', json_string)
        logger.debug(json_string)
        # write json string to file
        i = int(team.split()[-1]) - 1
        bestandsnaam = os.path.join(
            ni_dir, 'individueel_{0}_{1}.json'.format(
                team.lower().replace(' ', '_'), reeksnamen[i]))
        backup_vorig_bestand(bestandsnaam)
        with open(bestandsnaam, 'w') as f:
            f.write(json_string)

# https://docs.python.org/2/howto/urllib2.html
# http://www.diveintopython.net/http_web_services/etags.html
# https://docs.python.org/2/library/urllib2.html
# http://www.diveintopython3.net/http-web-services.html
# https://github.com/dustin/snippets/blob/master/python/net/http/fetch.py


def download_nieuwe_versie_bestand(bestandsnaam, verbose=False):
    url = url_locatie + bestandsnaam
    etag = read_etag(bestandsnaam)
    try:
        request = urllib2.Request(url)
        if etag:
            request.add_header('If-None-Match', etag)
        response = urllib2.urlopen(request)
        # no error raised, so the remote file changed
        # save remote file to disk
        logger.debug("downloading '{0}' ...".format(url))
        with open(bestandsnaam, "wb") as f:
            f.write(response.read())
        logger.debug("download okay")
        # store new etag value
        write_etag(bestandsnaam, response.info().getheader("ETag"))
    except urllib2.HTTPError as e:
        if e.code == 304:
            # geen nieuwe versie van bestand (code: 304)
            logger.info("HTTP Error: {0} {1}".format(e.code, url))
        else:
            logger.error("HTTP Error: {0} {1}".format(e.code, url))
    except urllib2.URLError as e:
        # remote server kan niet bereikt worden
        logger.error("URL Error: {0} {1}".format(e.reason, url))
    except IOError:
        logger.error(
            "bestand '{0}' kan niet geschreven worden".format(bestandsnaam))


def backup_vorig_bestand(bestand):
    if os.path.isfile(bestand):
        modifiedTime = os.path.getmtime(bestand)
        dt = datetime.datetime.fromtimestamp(
            modifiedTime).strftime("%Y-%m-%d_%H.%M.%S")
        (root, ext) = os.path.splitext(bestand)
        nieuwe_naam = root + '_' + dt + ext
        logger.debug("oude versie '{0}' hernoemen naar '{1}'".format(
            bestand, nieuwe_naam))
        os.rename(bestand, nieuwe_naam)


def read_etag(bestandsnaam):
    etag_bestandsnaam = bestandsnaam + ".etag"
    if os.path.isfile(etag_bestandsnaam):
        with open(etag_bestandsnaam, "r") as f:
            etag = f.read().strip()
        return etag
    else:
        logger.warning("ETag voor {0} niet gevonden".format(bestandsnaam))
        return None


def write_etag(bestandsnaam, etag):
    if etag:
        etag_bestandsnaam = bestandsnaam + ".etag"
        with open(etag_bestandsnaam, "w") as f:
            f.write(etag)
    else:
        logger.warning("ETag voor {0} niet gevonden".format(bestandsnaam))


def start_logging(verbose=False):
    # logging to console
    sh = logging.StreamHandler(sys.stdout)
    sh.setLevel(logging.DEBUG)
    formatter = logging.Formatter('[%(asctime)s] %(levelname)s\n%(message)s')
    sh.setFormatter(formatter)
    if os.access("/var/log/", os.W_OK):
        # logging to /var/log
        fh = logging.FileHandler(log_file)
    else:
        fh = logging.NullHandler()
    fh.setLevel(logging.INFO)
    formatter = logging.Formatter('[%(asctime)s] %(levelname)s - %(message)s')
    fh.setFormatter(formatter)
    # create Logger object
    logger = logging.getLogger(__name__)
    if verbose:
        logger.setLevel(logging.DEBUG)
    else:
        logger.setLevel(logging.INFO)
    logger.addHandler(sh)
    logger.addHandler(fh)
    return logger


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="update uitslagen interclubs op basis van data-bestanden kbsb")
    parser.add_argument("-p", "--ploegen", action="store_true",
                        help="update de ploeguitslagen en tussenstand")
    parser.add_argument("-i", "--individueel", action="store_true",
                        help="update de individuele uitslagen")
    parser.add_argument("-a", "--alles", action="store_true",
                        help="update alles")
    parser.add_argument("-v", "--verbose", action="store_true",
                        help="geef extra output in de console")
    args = parser.parse_args()
    logger = start_logging(args.verbose)
    if args.ploegen or args.alles:
        # openpyxl needs jdcal
        from openpyxl import load_workbook
        try:
            download_nieuwe_versie_bestand(xls_input, args.verbose)
            if os.path.isfile(xls_input):
                #maak_csv_ranking(args.verbose)
                maak_json_ploeg_uitslagen(args.verbose)
        finally:
            try:
                os.remove(xls_input)
            except OSError:
                pass
    if args.individueel or args.alles:
        try:
            download_nieuwe_versie_bestand(zip_input, args.verbose)
            if os.path.isfile(zip_input):
                with ZipFile(zip_input, 'r') as f:
                    f.extract(dbf_input)
                maak_json_individuele_uitslagen(args.verbose)
        finally:
            try:
                os.remove(zip_input)
                os.remove(dbf_input)
            except OSError:
                pass
    # geen argumenten opgegeven --> help tekst
    if not (args.ploegen or args.individueel or args.alles):
        parser.print_help()
