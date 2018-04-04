#!/usr/bin/env python
# -*- coding: utf-8 -*-

import argparse
from openpyxl import load_workbook
import csv
import json
import sys
from collections import OrderedDict
import re
import codecs


class generator:
    def __init__(self, seizoen, geen_kruistabel=False, verbosity=False):
        # definieer namen van bestanden en mappen
        self._ni_directory = "../_data/ni/" + seizoen + "/"
        self._geen_kruistabel = geen_kruistabel
        self._verbosity = verbosity
        self._pages_directory = "../_interclubs/"
        self._excel_filename = "individueel_eindstand_dworp_"
        self._excel_extensie = ".xlsx"
        self._csv_output = "eindstand_dworp_"
        self._json_output = "individueel_dworp_"
        self._html_output = "interclubs-"
        self._html_indiv_output = "interclubs-indiv-"
        # definieer namen van werkbladen
        self._ws_info = "Info"
        self._ws_ronden = ("R1", "R2", "R3", "R4", "R5",
                           "R6", "R7", "R8", "R9", "R10", "R11")
        self._ws_tabellen = "Ranking"
        # definieer cellen en ranges in excel-bestand
        self._cel_datum = "C1"
        self._cell_thuisploeg = ["C3", "C13", "C21", "C29"]
        self._cell_uitploeg = ["I3", "I13", "I21", "I29"]
        self._cell_thuisscore = ["E11", "E19", "E27", "E35"]
        self._cell_thuisscore_begin = ["E5", "E15", "E23", "E31"]
        self._cell_thuisscore_einde = ["E10", "E18", "E26", "E34"]
        self._cell_uitscore_begin = ["G5", "G15", "G23", "G31"]
        self._cell_uitscore_einde = ["G10", "G18", "G26", "G34"]
        self._cell_uitscore = ["G11", "G19", "G27", "G35"]
        self._cell_thuisgem = ["C11", "C19", "C27", "C35"]
        self._cell_uitgem = ["I11", "I19", "I27", "I35"]
        self._cell_indiv_begin = ["B5", "B15", "B23", "B31"]
        self._cell_indiv_einde = ["J10", "J18", "J26", "J34"]
        self._cel_tabel_begin = ["A3", "A17", "A31", "A45"]
        self._cel_tabel_einde = ["P15", "P29", "P43", "P57"]
        self._cel_tegenstanders_begin = ["B4", "B18", "B32", "B46"]
        self._cel_tegenstanders_eind = ["B15", "B29", "B43", "B57"]
        self._open_excel_bestand()
        self._lees_jaren()
        self._lees_reeksen()

    def _open_excel_bestand(self):
        for teams in ["1", "12", "123", "1234"]:
            seizoen = "_" + self._ni_directory[-5:-1]
            xls_input = self._ni_directory + self._excel_filename + \
                teams + seizoen + self._excel_extensie
            try:
                self._wb = load_workbook(xls_input, data_only=True)
                break
            except:
                if teams == "1234":
                    sys.exit(
                        "Fout! Er bestaat geen bestand in de vorm van: %s !" %
                        (self._ni_directory + self._excel_filename +
                         "***" + seizoen + self._excel_extensie))

    def _lees_jaren(self):
        self._jaren = [c.value for r in self._wb[self._ws_info]
                       ['B3':'B4'] for c in r]

    def _lees_reeksen(self):
        self._reeksen = [
            c.value for r in self._wb[self._ws_info]['B5':'B8'] for c in r]
        self._ploegen = range(len([r for r in self._reeksen if r]))

    def _verbose(self, info):
        if self._verbosity:
            if isinstance(info, unicode):
                info = info.encode('utf-8', 'ignore')
            print info

    def _html_interclubs(self, jaar1, jaar2, jaar12):
        html_string = (
            "---\n"
            "title: Nationale Interclubs {0} - {1}\n"
            "beginjaar: {0}\n"
            "---\n"
        ).format(jaar1, jaar2)
        self._schrijf_naar_bestand(
            self._pages_directory + self._html_output + jaar12 + ".html",
            html_string)

    def _schrijf_naar_bestand(self, bestandsnaam, s, altijd_schrijven=False):
        self._verbose(s)
        if s or altijd_schrijven:
            with codecs.open(bestandsnaam, 'w', encoding="utf-8") as f:
                f.write(s)

    def html(self):
        jaar1 = str(self._jaren[0])
        jaar2 = str(self._jaren[1])
        jaar12 = jaar1[2:] + jaar2[2:]
        self._html_interclubs(jaar1, jaar2, jaar12)

    def csv(self, ploeg=0):
        if ploeg:
            ploegen = [ploeg - 1]
        else:
            ploegen = self._ploegen
        ws = self._wb[self._ws_tabellen]
        for p in ploegen:
            cell_range = ws[self._cel_tabel_begin[p]:self._cel_tabel_einde[p]]
            rijen = []
            for row in cell_range:
                rij = []
                for cell in row:
                    if cell.value:
                        if rijen:
                            rij.append(unicode(cell.value).encode('utf-8'))
                        else:
                            # eerste rij lowercase
                            rij.append(
                                unicode(cell.value).encode('utf-8').lower())
                    elif cell.value == 0:
                        rij.append('0')
                    else:
                        rij.append('')
                rijen.append(rij)
            # strip laatste lege rijen uit de tabel
            while not rijen[-1][1]:
                del rijen[-1]
                for rij in rijen:
                    del rij[-3]
            # maak een gewone eindstand, geen kruistabel
            if self._geen_kruistabel:
                rijen = [rij[0:2] + rij[-2:] for rij in rijen]
            # verwijder lege tabel
            if rijen[-1][1] == 'team':
                rijen = []
            # schrijf csv-bestand
            bestandsnaam = self._ni_directory + self._csv_output + \
                str(p + 1) + "_" + self._reeksen[p].lower() + ".csv"
            self._verbose(rijen)
            if rijen:
                with open(bestandsnaam, 'wb') as csvfile:
                    csvwriter = csv.writer(csvfile, quoting=csv.QUOTE_NONE)
                    csvwriter.writerows(rijen)

    def json(self, ploeg=0):
        if ploeg:
            ploegen = [ploeg - 1]
        else:
            ploegen = self._ploegen
        for p in ploegen:
            ronde_lijst = []
            for r in range(1, 12):
                ws = self._wb["R" + str(r)]
                try:
                    datum = ws[self._cel_datum].value.strftime('%d-%m-%Y')
                except:
                    datum = None
                data = ["%d" % (r,), datum,
                        ws[self._cell_thuisploeg[p]].value,
                        ws[self._cell_uitploeg[p]].value,
                        ws[self._cell_thuisscore[p]].value,
                        ws[self._cell_uitscore[p]].value,
                        ws[self._cell_thuisgem[p]].value,
                        ws[self._cell_uitgem[p]].value]
                # afronden gemiddelde elo
                if isinstance(data[6], float):
                    data[6] = int(round(data[6]))
                if isinstance(data[7], float):
                    data[7] = int(round(data[7]))
                # verwijderen van None waarden uit 'data'
                data = [unicode(j).encode('utf-8')
                        for j in ["" if i is None else i for i in data]]
                ronde = OrderedDict(zip(
                    ("ronde", "datum", "thuis", "uit",
                     "tscore", "uscore", "tgem", "ugem"),
                    data))
                borden = []
                cell_range = ws[self._cell_indiv_begin[p]:
                                self._cell_indiv_einde[p]]
                for row in cell_range:
                    offset = (3, 5, 1, 7, 2, 8, 0, 6)
                    bord = ["" if j is None else j for j in [
                        row[i].value for i in offset]]
                    bord = [unicode(i).encode('utf-8') for i in bord]
                    # voeg alleen bord toe als er ergens iets is ingevuld in de
                    # rij
                    if any(bord):
                        borden.append(OrderedDict(zip(
                            ("tr", "ur", "tspeler", "uspeler",
                             "telo", "uelo", "tstam", "ustam"),
                            bord)))
                ronde["borden"] = borden
                # voeg alleen ronde toe als er ergens iets is ingevuld in excel
                if borden or any(data[2:]):
                    ronde_lijst.append(ronde)
            json_string = json.dumps(
                ronde_lijst, indent=2, separators=(',', ': '))
            # verwijder het teveel aan whitspace en newlines met behulp van
            # regular expressions
            r = re.compile(r'''
                (\{)\s+
                ("tr":\s.*,)\s+
                ("ur":\s.*,)\s+
                ("tspeler":\s.*,)\s+
                ("uspeler":\s.*,)\s+
                ("telo":\s.*,)\s+
                ("uelo":\s.*,)\s+
                ("tstam":\s.*,)\s+
                ("ustam":\s.*)\s+
                (\})
                ''', re.VERBOSE)
            json_string = r.sub(r'\1\2 \3 \4 \5 \6 \7 \8 \9\10', json_string)
            if json_string != '[]':
                self._schrijf_naar_bestand(
                    self._ni_directory + self._json_output +
                    str(p + 1) + "_" + self._reeksen[p].lower() + ".json",
                    json_string)

    def _lees_kruistabel_in(self):
        ws = self._wb[self._ws_tabellen]
        self._tegenstanders = []
        for p in self._ploegen:
            cell_range = ws[self._cel_tegenstanders_begin[p]:
                            self._cel_tegenstanders_eind[p]]
            namen = [cell.value.lower() for row in cell_range for cell in row
                     if isinstance(cell.value, unicode)]
            for i, n in enumerate(namen):
                if n.find("dworp") != -1:
                    dworp = i
                    rij_dworp = int(
                        self._cel_tegenstanders_begin[p][1:]) + dworp
                    uitslagen = [ws.cell(row=rij_dworp, column=(
                        3 + j)).value for j in range(len(namen))]
                    tegenstanders = zip(namen, uitslagen)
                    del tegenstanders[dworp]
                    break
            else:
                # dworp niet gevonden in de (waarschijnlijk lege) kruistabel
                tegenstanders = []
                self._validatie_string += \
                    u"Kruistabel: Dworp {0}: fout: ploeg van Dworp staat er " \
                    u"niet in\n".format(p + 1)
            self._tegenstanders.append(tegenstanders)

    def _vergelijk_individueel_met_kruistabel(self):
        for r in range(1, 12):
            ws = self._wb["R" + str(r)]
            for p in self._ploegen:
                try:
                    try:
                        if ws[self._cell_thuisploeg[p]].value.lower().find(
                                "dworp") == -1:
                            tegenstander = ws[
                                self._cell_thuisploeg[p]].value.lower()
                            score = int(ws[self._cell_uitscore[p]].value * 10)
                        else:
                            tegenstander = ws[
                                self._cell_uitploeg[p]].value.lower()
                            score = int(
                                ws[self._cell_thuisscore[p]].value * 10)
                    except TypeError:
                        # wanneer er geen som van de individuele resultaten is
                        # ingevuld krijgen we hier een TypeError
                        score = -9999
                    for n in self._tegenstanders[p]:
                        if n[0].find(tegenstander) != -1:
                            if tegenstander != n[0]:
                                # geen zuivere exacte match, bv. alleen
                                # ploegnummer in kruistabel
                                self._validatie_string += \
                                    u"Ronde {0}: Dworp {1}: fout: '{2}' " \
                                    u"niet identiek aan '{3}' in kruistabel" \
                                    u"\n".format(r, p + 1, tegenstander, n[0])
                            try:
                                if score != int(n[1] * 10):
                                    self._validatie_string += \
                                        u"Ronde {0}: Dworp {1}: fout: score " \
                                        u"tegen '{2}' niet identiek aan " \
                                        u"score in kruistabel\n".format(
                                            r, p + 1, tegenstander)
                            except TypeError:
                                self._validatie_string += \
                                    u"Ronde {0}: Dworp {1}: waarschuwing: " \
                                    u"score tegen '{2}' niet ingevuld in " \
                                    u"kruistabel\n".format(
                                        r, p + 1, tegenstander)
                            break
                    else:
                        # tegenstander niet gevonden, we proberen te zoeken met
                        # de ploegnaam zonder ploegnummer
                        try:
                            int(tegenstander.split()[-1])
                            verkorte_naam = u" ".join(
                                tegenstander.split()[:-1])
                        except ValueError:
                            verkorte_naam = tegenstander
                        for n in self._tegenstanders[p]:
                            if n[0].find(verkorte_naam) != -1:
                                self._validatie_string += \
                                    u"Ronde {0}: Dworp {1}: fout: '{2}' niet" \
                                    u" identiek aan '{3}' in kruistabel" \
                                    u"\n".format(r, p + 1, tegenstander, n[0])
                                try:
                                    if score != int(n[1] * 10):
                                        self._validatie_string += \
                                            u"Ronde {0}: Dworp {1}: fout: " \
                                            u"score tegen '{2}' niet " \
                                            u"identiek aan score in " \
                                            u"kruistabel\n".format(
                                                r, p + 1, tegenstander)
                                except TypeError:
                                    self._validatie_string += \
                                        u"Ronde {0}: Dworp {1}: waarschuwing" \
                                        u": score tegen '{2}' niet ingevuld " \
                                        u"in kruistabel\n".format(
                                            r, p + 1, tegenstander)
                                break
                        else:
                            # tegenstander nog steeds niet gevonden,
                            # waarschijnlijk een probleem met spelling
                            self._validatie_string += \
                                u"Ronde {0}: Dworp {1}: fout: '{2}' niet " \
                                u"gevonden in kruistabel (spelling?)\n".format(
                                    r, p + 1, tegenstander)
                            self._validatie_string += \
                                u"Ronde {0}: Dworp {1}: waarschuwing: score " \
                                u"tegen '{2}' is niet gecontroleerd in " \
                                u"kruistabel\n".format(r, p + 1, tegenstander)
                except AttributeError:
                    self._validatie_string += \
                        u"Ronde {0}: Dworp {1}: waarschuwing: geen ploegnaam " \
                        u"gevonden\n".format(r, p + 1)

    def _check_individuele_scores(self):
        for r in range(1, 12):
            ws = self._wb["R" + str(r)]
            for p in self._ploegen:
                cell_range = ws[self._cell_thuisscore_begin[p]:
                                self._cell_thuisscore_einde[p]]
                thuisscores = [cell.value for row in cell_range
                               for cell in row
                               if isinstance(cell.value, (int, long, float))]
                cell_range = ws[self._cell_uitscore_begin[p]:
                                self._cell_uitscore_einde[p]]
                uitscores = [cell.value for row in cell_range for cell in row
                             if isinstance(cell.value, (int, long, float))]
                if thuisscores or uitscores:
                    thuisscore = ws[self._cell_thuisscore[p]].value
                    uitscore = ws[self._cell_uitscore[p]].value
                    if len(thuisscores) == len(uitscores):
                        if sum(thuisscores) != thuisscore:
                            self._validatie_string += \
                                u"Ronde {0}: Dworp {1}: fout: som " \
                                u"thuisscores '{2}' niet gelijk aan totaal " \
                                u"'{3}'\n".format(
                                    r, p + 1, sum(thuisscores), thuisscore)
                        if sum(uitscores) != uitscore:
                            self._validatie_string += \
                                u"Ronde {0}: Dworp {1}: fout: som uitscores " \
                                u"'{2}' niet gelijk aan totaal '{3}'\n".format(
                                    r, p + 1, sum(uitscores), uitscore)
                        for b, (t, u) in enumerate(
                                zip(thuisscores, uitscores)):
                            if (t + u) != 1:
                                self._validatie_string += \
                                    u"Ronde {0}: Dworp {1}: Bord {2}: fout: " \
                                    u"som van bordscore niet gelijk aan 1" \
                                    u"\n".format(r, p + 1, b + 1)
                    else:
                        self._validatie_string += \
                            u"Ronde {0}: Dworp {1}: fout: individuele " \
                            u"uitslag niet volledig\n".format(r, p + 1)

    def valideer(self):
        self._validatie_string = u""
        self._check_individuele_scores()
        self._lees_kruistabel_in()
        self._vergelijk_individueel_met_kruistabel()
        self._schrijf_naar_bestand(
            self._ni_directory + "validatie.txt", self._validatie_string, True)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="maak webpagina van Eddy zijn nationale interclubs "
        "excel-bestanden")
    parser.add_argument("seizoen", help="geef het seizoen op, bv. 'ni9394'")
    parser.add_argument("-l", "--lint", action="store_true",
                        help="valideer gegevens in excel-bestand")
    parser.add_argument("-m", "--html", action="store_true",
                        help="maak nieuwe jekyll html-pagina")
    parser.add_argument("-j", "--json", type=int, choices=[0, 1, 2, 3, 4],
                        help="maak json-bestand met individuele resultaten "
                        "van ploegnummer")
    parser.add_argument("-c", "--csv", type=int, choices=[0, 1, 2, 3, 4],
                        help="maak csv-bestand met eindstand van ploegnummer")
    parser.add_argument("-s", "--summier", action="store_true",
                        help="eindstand bevat enkel totalen, geen onderlinge resultaten")
    parser.add_argument("-a", "--alles", action="store_true",
                        help="doe alles in eens")
    parser.add_argument("-v", "--verbose", action="store_true",
                        help="geef ook output in de console")
    args = parser.parse_args()
    g = generator(args.seizoen, args.summier, args.verbose)
    if args.alles:
        print "doe alles"
        g.valideer()
        g.csv()
        g.json()
        g.html()
    else:
        if args.lint:
            print("Valideer input in excel-bestand.")
            g.valideer()
        if args.csv:
            print("Maak csv voor Dworp %d" % args.csv)
            g.csv(args.csv)
        elif args.csv == 0:
            print("Maak csv-bestand voor alle ploegen.")
            g.csv()
        if args.json:
            print("Maak json-bestond voor Dworp %d." % args.json)
            g.json(args.json)
        elif args.json == 0:
            print("Maak json-bestand voor alle ploegen.")
            g.json()
        if args.html:
            print("Maak html-bestanden.")
            g.html()
        # default
        if not any((args.lint, args.csv, args.json, args.html, args.json == 0,
                    args.csv == 0)):
            print("Valideer input in excel-bestand (default).")
            g.valideer()
